#!/usr/bin/env python3
"""
enrich_by_siret.py — Batch INSEE Sirene API v3.11 lookup by SIRET only.

Reads `FR_SIRET`, calls GET /siret/{siret}, flattens JSON as-is into columns.
No name search, no field hardcoding (schema from API responses).

Auth (first match wins):
  1) INSEE_TOKEN — OAuth2 Bearer (optional INSEE_SIRENE_BASE_URL, default entreprises URL)
  2) same as enrich_suppliers — VITE_INSEE_API_KEY … VITE_INSEE_API_KEY10 + rotator
     (VITE_API_BASE_URL default https://api.insee.fr/api-sirene/3.11)

Usage:
    python enrich_by_siret.py [input.xlsx] [--output enriched_by_siret.xlsx] [--workers N]
    python enrich_by_siret.py input.xlsx --limit 5 --debug-json mapping.json
    python enrich_by_siret.py input.xlsx --integration-keys   # same keys as enrich_suppliers (ignores INSEE_TOKEN)

Output Excel follows excel_skill.md: plain ranges (no tables), freeze row 1, auto_filter,
column tints (FR_* / AI_* / INSEE payload columns), INSEE_STATUS cell colors (no zebra rows).

Dependencies:
    pip install pandas openpyxl requests python-dotenv
"""
from __future__ import annotations

import argparse
from copy import copy
import json
import logging
import math
import os
import re
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
import requests
from dotenv import load_dotenv
from insee_key_rotator import active_keys, get_next_insee_key, throttle_insee
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

load_dotenv()

INSEE_BASE = "https://api.insee.fr/entreprises/sirene/V3.11"
DEFAULT_API_SIRENE_BASE = "https://api.insee.fr/api-sirene/3.11"
COL_SIRET = "FR_SIRET"
COL_STATUS = "INSEE_STATUS"
LOG_FILE = "siret_errors.log"
SWAGGER_SPEC_FILE = "api-spec.json"

# Coordinate rate limiting + key rotation safely while still allowing concurrent requests.
_INSEE_RATE_LOCK = threading.Lock()
_last_bearer_gap_ts = 0.0


@dataclass
class InseeAuth:
    use_bearer: bool
    bearer_token: str
    base_url: str


def _env_float(name: str, default: float) -> float:
    v = os.environ.get(name)
    if v is None or v == "":
        return default
    try:
        return float(v)
    except ValueError:
        return default


def _env_int(name: str, default: int) -> int:
    v = os.environ.get(name)
    if v is None or v == "":
        return default
    try:
        return int(v)
    except ValueError:
        return default


# ~30 req/min default; override with INSEE_MIN_INTERVAL_SEC (e.g. 0.2 for 300/min)
INSEE_MIN_INTERVAL_SEC = _env_float("INSEE_MIN_INTERVAL_SEC", 60.0 / 30.0)
MAX_WORKERS = _env_int("INSEE_MAX_WORKERS", 8)


def _throttle_bearer_gap() -> None:
    """Extra spacing when using Bearer (INSEE_MIN_INTERVAL_SEC); runs inside _INSEE_FETCH_LOCK."""
    global _last_bearer_gap_ts
    now = time.monotonic()
    gap = INSEE_MIN_INTERVAL_SEC - (now - _last_bearer_gap_ts)
    if gap > 0:
        time.sleep(gap)
    _last_bearer_gap_ts = time.monotonic()


_insee_auth: InseeAuth | None = None


def _siret_digits(raw: Any) -> str:
    if _is_empty_siret_cell(raw):
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    return re.sub(r"\D", "", s)


def _maybe_pad_siret_to_14(siret_digits: str) -> str | None:
    """
    Retry helper for common Excel issue where leading zeros are dropped.
    Only pad when it is plausibly missing 1-2 leading zeros (12-13 digits).
    """
    if not siret_digits:
        return None
    if len(siret_digits) in (12, 13):
        return siret_digits.zfill(14)
    return None


def _is_empty_siret_cell(v: Any) -> bool:
    if v is None:
        return True
    try:
        if pd.isna(v):
            return True
    except TypeError:
        pass
    if isinstance(v, str) and not v.strip():
        return True
    return False


def flatten_json(obj: Any, prefix: str = "") -> dict[str, Any]:
    """
    Flatten nested dicts with '_' separator. Keys match INSEE JSON leaf paths.
    Lists of dicts: index segments (e.g. periodesEtablissement_0_key).
    Lists of scalars / mixed: JSON string in one column.
    """
    out: dict[str, Any] = {}
    if isinstance(obj, dict):
        if not obj:
            if prefix:
                out[prefix] = ""
            return out
        for k, v in obj.items():
            p = f"{prefix}_{k}" if prefix else str(k)
            out.update(flatten_json(v, p))
        return out
    if isinstance(obj, list):
        if not obj:
            if prefix:
                out[prefix] = ""
            return out
        if all(isinstance(x, dict) for x in obj):
            for i, item in enumerate(obj):
                p = f"{prefix}_{i}" if prefix else str(i)
                out.update(flatten_json(item, p))
            return out
        if prefix:
            out[prefix] = json.dumps(obj, ensure_ascii=False)
        return out
    if prefix:
        out[prefix] = obj
    return out


def _cell_value(v: Any) -> Any:
    """Excel-safe scalar; JSON-like values serialized consistently for flat mapping."""
    if v is None:
        return ""
    if isinstance(v, float):
        if pd.isna(v) or (isinstance(v, float) and math.isnan(v)):
            return ""
        return v
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, bool):
        return v
    if isinstance(v, int):
        return v
    if isinstance(v, str):
        return v
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False, separators=(",", ":"))
    return str(v)


def _fetch_siret(
    siret: str,
) -> tuple[int, dict[str, Any] | None, str]:
    """
    Returns (http_status, flattened_json_or_none, error_message_for_log).
    """
    auth = _insee_auth
    if auth is None:
        return 0, None, "auth not configured"

    try:
        with _INSEE_RATE_LOCK:
            if auth.use_bearer:
                _throttle_bearer_gap()
                headers = {
                    "Authorization": f"Bearer {auth.bearer_token}",
                    "Accept": "application/json",
                }
            else:
                throttle_insee()
                key = get_next_insee_key()
                headers = {
                    "X-INSEE-Api-Key-Integration": key,
                    "Accept": "application/json",
                }
            url = f"{auth.base_url.rstrip('/')}/siret/{siret}"
        r = requests.get(url, headers=headers, timeout=60)
        status = r.status_code
        if status == 200:
            try:
                data = r.json()
            except Exception as e:
                return status, None, f"JSON parse error: {e}"
            flat = flatten_json(data)
            return status, flat, ""
        if status == 404:
            return status, None, "NOT_FOUND"
        try:
            body = r.text[:500]
        except Exception:
            body = ""
        return status, None, body or r.reason or "HTTP error"
    except requests.RequestException as e:
        return 0, None, str(e)


_schema_lock = threading.Lock()
_insee_leaf_keys: list[str] = []


def _set_schema_from_first_success(flat: dict[str, Any]) -> None:
    """First successful 200 response defines all INSEE column names (sorted keys)."""
    global _insee_leaf_keys
    with _schema_lock:
        if _insee_leaf_keys:
            return
        _insee_leaf_keys = sorted(flat.keys())


def _setup_error_logger(path: Path) -> logging.Logger:
    log = logging.getLogger("siret_errors")
    log.handlers.clear()
    log.setLevel(logging.INFO)
    fh = logging.FileHandler(path, encoding="utf-8")
    fh.setFormatter(logging.Formatter("%(asctime)s\t%(message)s"))
    log.addHandler(fh)
    return log


def _worker(
    row_index: int,
    siret: str,
    err_logger: logging.Logger,
) -> tuple[int, str, dict[str, Any] | None, int]:
    """
    Returns (row_index, insee_status, flat_dict_or_none, http_status).
    """
    try:
        http_st, flat, err_msg = _fetch_siret(siret)
        if http_st == 200 and flat is not None:
            _set_schema_from_first_success(flat)
            return row_index, "OK", flat, http_st
        if http_st == 404:
            status = "NOT_FOUND"
        elif http_st == 0:
            status = "ERROR"
        else:
            status = str(http_st)
        err_logger.info(
            f"{siret}\t{http_st if http_st else 'N/A'}\t{err_msg or status}",
        )
        return row_index, status, None, http_st
    except Exception as e:
        err_logger.info(f"{siret}\tEXC\t{e!s}")
        return row_index, "ERROR", None, 0


def _approximate_sheet_column_widths(ws, max_width: float = 60.0) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            max_len = max(max_len, min(len(s), 80))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), max_width)


# ── excel_skill.md: plain ranges, no tables; freeze + filter + column tint + status cells ──
FILL_HEADER = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_COL_FR = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_COL_AI = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
FILL_COL_INSEE_BODY = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FONT_DEFAULT = Font(color="000000")
# INSEE_STATUS — map to excel_skill FR_Status palette where applicable
FILL_ST_OK = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
FONT_ST_OK = Font(color="FFFFFF", bold=True)
FILL_ST_ERR = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
FONT_ST_ERR = Font(color="FFFFFF", bold=True)
FONT_SIRET_ERROR = Font(color="FF0000")
COMMENT_AUTHOR = "INSEE Swagger"


def _base_column_fill(col_name: str, insee_leaf_set: set[str]) -> PatternFill | None:
    if col_name == COL_STATUS:
        return None
    if col_name in insee_leaf_set:
        return FILL_COL_INSEE_BODY
    if col_name.startswith("FR_"):
        return FILL_COL_FR
    if col_name.startswith("AI_"):
        return FILL_COL_AI
    return None


def _insee_status_cell_style(st: str) -> tuple[PatternFill, Font]:
    if st == "OK":
        return FILL_ST_OK, FONT_ST_OK
    if st:
        return FILL_ST_ERR, FONT_ST_ERR
    return FILL_ST_ERR, FONT_ST_ERR


_swagger_desc_cache: dict[str, str] | None = None


def _camel_to_snake(name: str) -> str:
    s = re.sub(r"(.)([A-Z][a-z]+)", r"\1_\2", name)
    s = re.sub(r"([a-z0-9])([A-Z])", r"\1_\2", s)
    return s.lower()


def _description_key(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", _camel_to_snake(name).lower())


def _repair_mojibake(text: str) -> str:
    if not any(marker in text for marker in ("Ã", "Â", "Ă")):
        return text
    try:
        return text.encode("latin1").decode("utf-8")
    except UnicodeError:
        return text


def _strip_known_suffixes(name: str) -> str:
    for suffix in (
        "_etablissement",
        "_unite_legale",
        "_unitelegale",
        "_entreprise",
    ):
        if name.endswith(suffix):
            return name[: -len(suffix)]
    return name


def _load_swagger_descriptions() -> dict[str, str]:
    global _swagger_desc_cache
    if _swagger_desc_cache is not None:
        return _swagger_desc_cache

    descriptions: dict[str, str] = {}
    spec_path = Path(__file__).resolve().parent / SWAGGER_SPEC_FILE
    if not spec_path.is_file():
        _swagger_desc_cache = descriptions
        return descriptions

    try:
        spec = json.loads(spec_path.read_text(encoding="utf-8-sig"))
    except Exception:
        _swagger_desc_cache = descriptions
        return descriptions

    schemas = spec.get("components", {}).get("schemas", {})
    for schema_name, schema in schemas.items():
        props = schema.get("properties", {})
        if not isinstance(props, dict):
            continue
        for prop_name, prop in props.items():
            if not isinstance(prop, dict):
                continue
            desc = prop.get("description")
            if not desc:
                continue
            clean_desc = _repair_mojibake(str(desc).strip())
            if not clean_desc:
                continue
            candidates = {
                prop_name,
                _camel_to_snake(prop_name),
                _strip_known_suffixes(_camel_to_snake(prop_name)),
                f"{schema_name}_{prop_name}",
                f"{schema_name}_{_camel_to_snake(prop_name)}",
            }
            for candidate in candidates:
                descriptions.setdefault(_description_key(candidate), clean_desc)

    _swagger_desc_cache = descriptions
    return descriptions


def _readable_column_label(col_name: str) -> str:
    parts = [p for p in col_name.split("_") if p and not p.isdigit()]
    if parts and parts[0] in {"etablissement", "header"}:
        parts = parts[1:]
    text = " ".join(_camel_to_snake(p).replace("_", " ") for p in parts)
    return re.sub(r"\s+", " ", text).strip().capitalize() or col_name


def _description_candidates_for_column(col_name: str) -> list[str]:
    parts = [p for p in col_name.split("_") if p and not p.isdigit()]
    leaf = parts[-1] if parts else col_name
    leaf_snake = _camel_to_snake(leaf)
    full_snake = _camel_to_snake(col_name)
    candidates = [
        col_name,
        full_snake,
        leaf,
        leaf_snake,
        _strip_known_suffixes(leaf_snake),
        _strip_known_suffixes(full_snake),
    ]
    if leaf_snake.startswith("changement_"):
        candidates.append(leaf_snake.removeprefix("changement_"))
    return candidates


def _header_description(col_name: str) -> str:
    fixed = {
        COL_STATUS: "Statut du lookup INSEE pour la ligne: OK, NOT_FOUND, INVALID_SIRET ou erreur HTTP/API.",
        "header_statut": "Statut HTTP renvoyé par l'API INSEE pour cette réponse.",
        "header_message": "Message renvoyé par l'API INSEE pour cette réponse.",
    }
    if col_name in fixed:
        return fixed[col_name]

    descriptions = _load_swagger_descriptions()
    for candidate in _description_candidates_for_column(col_name):
        desc = descriptions.get(_description_key(candidate))
        if desc:
            return f"{desc}\n\nNom technique: {col_name}"

    return f"{_readable_column_label(col_name)}.\n\nNom technique: {col_name}"


def _apply_header_comments(ws, column_map: dict[str, int]) -> None:
    for col_name, col_idx in column_map.items():
        desc = _header_description(col_name)
        if len(desc) > 900:
            desc = desc[:897].rstrip() + "..."
        ws.cell(row=1, column=col_idx).comment = Comment(desc, COMMENT_AUTHOR)


def write_workbook_excel_skill(
    out_path: Path,
    out_columns: list[str],
    out_rows: list[dict[str, Any]],
    insee_leaf_keys: list[str],
) -> None:
    """
    excel_skill.md: no ws.add_table(); freeze A2; auto_filter; column tints;
    INSEE_STATUS uses skill-style status colors (cells, not full-row zebra).
    """
    insee_leaf_set = set(insee_leaf_keys)
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    header_font = Font(bold=True)

    for c_idx, col_name in enumerate(out_columns, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.font = header_font
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    _apply_header_comments(
        ws,
        {
            col_name: idx
            for idx, col_name in enumerate(out_columns, start=1)
            if col_name == COL_STATUS or col_name in insee_leaf_set
        },
    )

    for r_idx, row in enumerate(out_rows, start=2):
        st = str(row.get(COL_STATUS, "") or "")
        for c_idx, col_name in enumerate(out_columns, start=1):
            val = row.get(col_name, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=_cell_value(val))
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if col_name == COL_STATUS:
                if st:
                    fill, font = _insee_status_cell_style(st)
                    cell.fill = fill
                    cell.font = font
                else:
                    cell.font = FONT_DEFAULT
                continue

            base = _base_column_fill(col_name, insee_leaf_set)
            if base:
                cell.fill = base
            cell.font = FONT_DEFAULT

            if col_name == COL_SIRET and st and st != "OK":
                cell.font = FONT_SIRET_ERROR

    last_row = max(1, len(out_rows) + 1)
    last_col = max(1, len(out_columns))
    if last_row >= 2:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"

    _approximate_sheet_column_widths(ws)
    wb.save(out_path)


def _sheet_from_workbook(wb, sheet_index: int, sheet_name: str | None):
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet {sheet_name!r} not found. Sheets: {wb.sheetnames}")
        return wb[sheet_name]
    if sheet_index < 0 or sheet_index >= len(wb.worksheets):
        raise ValueError(f"Sheet index {sheet_index} out of range. Sheets: {wb.sheetnames}")
    return wb.worksheets[sheet_index]


def _build_header_map(ws, header_row: int = 1) -> dict[str, int]:
    """Map header text -> column index (1-based). Trims strings; ignores blanks."""
    m: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        if s not in m:
            m[s] = c
    return m


def _last_header_col(ws, header_row: int = 1) -> int:
    for c in range(ws.max_column, 0, -1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return c
    return 1


def _ensure_columns_preserve(ws, col_names: list[str], header_row: int = 1) -> dict[str, int]:
    """
    Ensure each column exists (by header name) and return mapping name->col_index.
    New columns are appended after the last header column and styled by copying the
    last existing header cell style to preserve the workbook look.
    """
    header_map = _build_header_map(ws, header_row=header_row)
    last_col = _last_header_col(ws, header_row=header_row)
    template = ws.cell(row=header_row, column=last_col)

    out: dict[str, int] = {}
    for name in col_names:
        if name in header_map:
            out[name] = header_map[name]
            continue
        last_col += 1
        cell = ws.cell(row=header_row, column=last_col, value=name)
        if template.has_style:
            cell._style = copy(template._style)
        out[name] = last_col
        header_map[name] = last_col
    return out


def write_workbook_preserve_xlsx(
    in_path: Path,
    out_path: Path,
    sheet_index: int,
    sheet_name: str | None,
    corrected_siret_by_row: dict[int, str],
    results_by_excel_row: dict[int, tuple[str, dict[str, Any] | None]],
    skipped_excel_rows: set[int],
    invalid_excel_rows: set[int],
    invalid_values_by_row: dict[int, str],
    insee_leaf_keys: list[str],
) -> None:
    """
    Preserve the original workbook (comments, formatting, merges, etc.) and only
    append/fill the INSEE columns on the selected sheet.
    """
    wb = load_workbook(in_path)
    ws = _sheet_from_workbook(wb, sheet_index=sheet_index, sheet_name=sheet_name)

    header_map = _build_header_map(ws, header_row=1)
    if COL_SIRET not in header_map:
        raise ValueError(f"Column {COL_SIRET!r} not found in header row 1.")
    siret_col = header_map[COL_SIRET]

    last_existing_col = _last_header_col(ws, header_row=1)
    cols_needed = [COL_STATUS] + insee_leaf_keys
    out_col_map = _ensure_columns_preserve(ws, cols_needed, header_row=1)
    _apply_header_comments(ws, out_col_map)

    for excel_row in range(2, ws.max_row + 1):
        if (
            excel_row not in results_by_excel_row
            and excel_row not in skipped_excel_rows
            and excel_row not in invalid_excel_rows
        ):
            continue

        tmpl = ws.cell(row=excel_row, column=min(last_existing_col, ws.max_column))
        tmpl_style = copy(tmpl._style) if tmpl.has_style else None

        # If we successfully recovered the SIRET by padding zeros, persist the corrected value.
        if excel_row in corrected_siret_by_row:
            ws.cell(row=excel_row, column=siret_col, value=corrected_siret_by_row[excel_row])

        if excel_row in skipped_excel_rows:
            ws.cell(row=excel_row, column=out_col_map[COL_STATUS], value="")
            for k in insee_leaf_keys:
                ws.cell(row=excel_row, column=out_col_map[k], value="")
        elif excel_row in invalid_excel_rows:
            ws.cell(
                row=excel_row,
                column=siret_col,
                value=invalid_values_by_row.get(excel_row, ""),
            )
            ws.cell(row=excel_row, column=out_col_map[COL_STATUS], value="INVALID_SIRET")
            for k in insee_leaf_keys:
                ws.cell(row=excel_row, column=out_col_map[k], value="")
        else:
            st, flat = results_by_excel_row.get(excel_row, ("", None))
            ws.cell(row=excel_row, column=out_col_map[COL_STATUS], value=st)
            for k in insee_leaf_keys:
                v = ""
                if flat and k in flat:
                    v = _cell_value(flat.get(k))
                ws.cell(row=excel_row, column=out_col_map[k], value=v)

        if tmpl_style is not None:
            ws.cell(row=excel_row, column=out_col_map[COL_STATUS])._style = copy(tmpl_style)
            for k in insee_leaf_keys:
                ws.cell(row=excel_row, column=out_col_map[k])._style = copy(tmpl_style)

    wb.save(out_path)


def _dump_debug_json(
    path: Path,
    df: pd.DataFrame,
    results: dict[int, tuple[str, dict[str, Any] | None]],
    skipped_indices: set[int],
    insee_leaf_keys: list[str],
) -> None:
    """Dump per-row API mapping + full flat dict for serialization inspection."""
    rows_out: list[dict[str, Any]] = []
    for i in range(len(df)):
        raw = df.iloc[i][COL_SIRET]
        siret = _siret_digits(raw)
        entry: dict[str, Any] = {
            "row_index": i,
            "fr_siret_raw": None if _is_empty_siret_cell(raw) else _cell_value(raw),
            "siret_digits": siret or None,
            "skipped_empty": i in skipped_indices,
        }
        if i in results:
            st, flat = results[i]
            entry["insee_status"] = st
            entry["flat_row"] = flat if flat is not None else {}
            entry["flat_keys_sorted"] = sorted(flat.keys()) if flat else []
        else:
            entry["insee_status"] = None
            entry["flat_row"] = {}
            entry["flat_keys_sorted"] = []
        rows_out.append(entry)
    payload = {
        "schema_insee_leaf_keys": insee_leaf_keys,
        "rows": rows_out,
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    ap = argparse.ArgumentParser(description="Enrich Excel by INSEE SIRET (Sirene API v3.11)")
    ap.add_argument(
        "input_xlsx",
        nargs="?",
        default="corrected frensh siret.xlsx",
        help="Input Excel path (default: corrected frensh siret.xlsx)",
    )
    ap.add_argument(
        "-o",
        "--output",
        default="enriched_by_siret.xlsx",
        help="Output Excel path (default: enriched_by_siret.xlsx)",
    )
    ap.add_argument(
        "--workers",
        type=int,
        default=MAX_WORKERS,
        help=f"ThreadPoolExecutor max workers (default: {MAX_WORKERS} or INSEE_MAX_WORKERS)",
    )
    ap.add_argument(
        "--sheet",
        type=int,
        default=0,
        help="Sheet index (0-based) or use --sheet-name",
    )
    ap.add_argument("--sheet-name", default=None, help="Sheet name (overrides --sheet)")
    ap.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Process only the first N rows of the sheet (0 = all)",
    )
    ap.add_argument(
        "--debug-json",
        default=None,
        metavar="PATH",
        help="Write mapping + full flat JSON per row for serialization checks",
    )
    ap.add_argument(
        "--integration-keys",
        action="store_true",
        help="Use VITE_INSEE_API_KEY rotator + api-sirene base (same as enrich_suppliers); ignores INSEE_TOKEN",
    )
    ap.add_argument(
        "--preserve",
        action="store_true",
        help="Preserve the original Excel workbook (comments/formatting) and only append new columns",
    )
    args = ap.parse_args()

    global _insee_auth, _insee_leaf_keys
    _insee_leaf_keys = []
    bearer = (os.environ.get("INSEE_TOKEN") or "").strip()
    keys = active_keys()
    if args.integration_keys and keys:
        base = (os.environ.get("VITE_API_BASE_URL") or DEFAULT_API_SIRENE_BASE).strip()
        _insee_auth = InseeAuth(use_bearer=False, bearer_token="", base_url=base)
        print(
            f"INSEE auth: X-INSEE-Api-Key-Integration ({len(keys)} key(s), base {base}) [forced]",
            file=sys.stderr,
        )
    elif bearer:
        base = (os.environ.get("INSEE_SIRENE_BASE_URL") or INSEE_BASE).strip()
        _insee_auth = InseeAuth(use_bearer=True, bearer_token=bearer, base_url=base)
        print(f"INSEE auth: Bearer (base {base})", file=sys.stderr)
    elif keys:
        base = (os.environ.get("VITE_API_BASE_URL") or DEFAULT_API_SIRENE_BASE).strip()
        _insee_auth = InseeAuth(use_bearer=False, bearer_token="", base_url=base)
        print(
            f"INSEE auth: X-INSEE-Api-Key-Integration ({len(keys)} key(s), base {base})",
            file=sys.stderr,
        )
    else:
        if args.integration_keys and not keys:
            print(
                "❌ --integration-keys requires at least one VITE_INSEE_API_KEY in .env",
                file=sys.stderr,
            )
        else:
            print(
                "❌ No INSEE credentials: set INSEE_TOKEN or at least VITE_INSEE_API_KEY in .env",
                file=sys.stderr,
            )
        sys.exit(1)

    in_path = Path(args.input_xlsx)
    if not in_path.is_file():
        print(f"❌ Missing input file: {in_path}", file=sys.stderr)
        sys.exit(1)

    out_path = Path(args.output)

    # Preserve mode: keep the original workbook (comments/formatting) and only append columns.
    # Implemented as an early-exit path to keep the existing behavior unchanged.
    if args.preserve:
        err_log_path = Path(__file__).resolve().parent / LOG_FILE
        err_logger = _setup_error_logger(err_log_path)

        _insee_leaf_keys = []

        wb = load_workbook(in_path)
        ws = _sheet_from_workbook(wb, sheet_index=args.sheet, sheet_name=args.sheet_name)
        header_map = _build_header_map(ws, header_row=1)
        if COL_SIRET not in header_map:
            print(f"❌ Column {COL_SIRET!r} not found in header row 1.", file=sys.stderr)
            sys.exit(1)
        siret_col = header_map[COL_SIRET]

        max_row = ws.max_row
        if args.limit and args.limit > 0:
            max_row = min(max_row, 1 + int(args.limit))

        results_xl: dict[int, tuple[str, dict[str, Any] | None]] = {}
        skipped_rows: set[int] = set()
        invalid_rows: set[int] = set()
        invalid_values_by_row: dict[int, str] = {}
        corrected_siret_by_row: dict[int, str] = {}
        to_fetch: list[tuple[int, str]] = []
        to_fetch_padded: list[tuple[int, str]] = []

        for excel_row in range(2, max_row + 1):
            raw = ws.cell(row=excel_row, column=siret_col).value
            if _is_empty_siret_cell(raw):
                skipped_rows.add(excel_row)
                continue
            siret = _siret_digits(raw)
            if len(siret) != 14:
                invalid_rows.add(excel_row)
                invalid_values_by_row[excel_row] = "" if _is_empty_siret_cell(raw) else str(raw)
                results_xl[excel_row] = ("INVALID_SIRET", None)
                err_logger.info(f"{raw}\tINVALID_SIRET\tExpected 14 digits after cleanup")
                padded = _maybe_pad_siret_to_14(siret)
                if padded:
                    to_fetch_padded.append((excel_row, padded))
                continue
            to_fetch.append((excel_row, siret))

        if to_fetch:
            with ThreadPoolExecutor(max_workers=max(1, args.workers)) as ex:
                futures = []
                for row_idx, siret in to_fetch:
                    futures.append(ex.submit(_worker, row_idx, siret, err_logger))
                for fut in as_completed(futures):
                    try:
                        row_index, st, flat, _http = fut.result()
                        results_xl[row_index] = (st, flat)
                    except Exception as e:
                        err_logger.info(f"future\tEXC\t{e!s}")

        # Second round: for rows with 12-13 digits, try padding leading zeros to 14 and refetch.
        # This is intentionally limited to error rows only (not the full list).
        if to_fetch_padded:
            padded_map = {row_idx: padded_siret for row_idx, padded_siret in to_fetch_padded}
            with ThreadPoolExecutor(max_workers=max(1, args.workers)) as ex:
                futures = []
                for row_idx, padded_siret in to_fetch_padded:
                    futures.append(ex.submit(_worker, row_idx, padded_siret, err_logger))
                for fut in as_completed(futures):
                    try:
                        row_index, st, flat, _http = fut.result()
                        if st == "OK" and flat is not None:
                            results_xl[row_index] = (st, flat)
                            corrected_siret_by_row[row_index] = padded_map.get(row_index, "")
                            invalid_rows.discard(row_index)
                    except Exception as e:
                        err_logger.info(f"future2\tEXC\t{e!s}")

        try:
            write_workbook_preserve_xlsx(
                in_path=in_path,
                out_path=out_path,
                sheet_index=args.sheet,
                sheet_name=args.sheet_name,
                corrected_siret_by_row=corrected_siret_by_row,
                results_by_excel_row=results_xl,
                skipped_excel_rows=skipped_rows,
                invalid_excel_rows=invalid_rows,
                invalid_values_by_row=invalid_values_by_row,
                insee_leaf_keys=_insee_leaf_keys,
            )
        except Exception as e:
            print(f"❌ Preserve-write failed: {e}", file=sys.stderr)
            sys.exit(1)

        n_total = max(0, max_row - 1)
        n_skipped = len(skipped_rows)
        n_found = sum(1 for _r, (st, _flat) in results_xl.items() if st == "OK")
        n_not_found = sum(1 for _r, (st, _flat) in results_xl.items() if st != "OK")

        lines = [
            f"Total rows: {n_total}",
            f"Found (OK): {n_found}",
            f"Not found / API error: {n_not_found}",
            f"Skipped (empty FR_SIRET): {n_skipped}",
            f"Output: {out_path.resolve()}",
            f"Error log: {err_log_path.resolve()}",
        ]
        print("\n".join(lines))
        return

    read_kw: dict[str, Any] = {"engine": "openpyxl", "dtype": {COL_SIRET: "string"}}
    if args.sheet_name:
        df = pd.read_excel(in_path, sheet_name=args.sheet_name, **read_kw)
    else:
        df = pd.read_excel(in_path, sheet_name=args.sheet, **read_kw)

    if COL_SIRET not in df.columns:
        print(f"❌ Column {COL_SIRET!r} not found. Columns: {list(df.columns)}", file=sys.stderr)
        sys.exit(1)

    if args.limit and args.limit > 0:
        df = df.head(args.limit).copy()

    orig_cols = list(df.columns)
    n_total = len(df)

    # Per-row results: row_index -> (status, flat_dict|None)
    results: dict[int, tuple[str, dict[str, Any] | None]] = {}
    skipped_indices: set[int] = set()

    to_fetch: list[tuple[int, str]] = []
    to_fetch_padded: list[tuple[int, str]] = []
    invalid_indices: set[int] = set()
    invalid_values: dict[int, str] = {}
    corrected_siret_by_index: dict[int, str] = {}
    err_log_path = Path(__file__).resolve().parent / LOG_FILE
    err_logger = _setup_error_logger(err_log_path)

    for i in range(n_total):
        raw = df.iloc[i][COL_SIRET]
        if _is_empty_siret_cell(raw):
            skipped_indices.add(i)
            continue
        siret = _siret_digits(raw)
        if len(siret) != 14:
            invalid_indices.add(i)
            invalid_values[i] = "" if _is_empty_siret_cell(raw) else str(raw)
            results[i] = ("INVALID_SIRET", None)
            err_logger.info(f"{raw}\tINVALID_SIRET\tExpected 14 digits after cleanup")
            padded = _maybe_pad_siret_to_14(siret)
            if padded:
                to_fetch_padded.append((i, padded))
            continue
        to_fetch.append((i, siret))

    _insee_leaf_keys = []

    if to_fetch:
        with ThreadPoolExecutor(max_workers=max(1, args.workers)) as ex:
            futures = []
            for row_idx, siret in to_fetch:
                futures.append(
                    ex.submit(_worker, row_idx, siret, err_logger),
                )
            for fut in as_completed(futures):
                try:
                    row_index, st, flat, _http = fut.result()
                    results[row_index] = (st, flat)
                except Exception as e:
                    err_logger.info(f"future\tEXC\t{e!s}")

    # Second round: only for rows that were invalid due to 12-13 digits, try left-padding zeros.
    if to_fetch_padded:
        padded_map = {row_idx: padded_siret for row_idx, padded_siret in to_fetch_padded}
        with ThreadPoolExecutor(max_workers=max(1, args.workers)) as ex:
            futures = []
            for row_idx, padded_siret in to_fetch_padded:
                futures.append(ex.submit(_worker, row_idx, padded_siret, err_logger))
            for fut in as_completed(futures):
                try:
                    row_index, st, flat, _http = fut.result()
                    if st == "OK" and flat is not None:
                        results[row_index] = (st, flat)
                        corrected_siret_by_index[row_index] = padded_map.get(row_index, "")
                        invalid_indices.discard(row_index)
                except Exception as e:
                    err_logger.info(f"future2\tEXC\t{e!s}")

    # Build output rows as list of dicts
    out_rows: list[dict[str, Any]] = []
    for i in range(n_total):
        row_dict: dict[str, Any] = {c: df.iloc[i][c] for c in orig_cols}
        if i in corrected_siret_by_index:
            row_dict[COL_SIRET] = corrected_siret_by_index[i]
        if i in skipped_indices:
            row_dict[COL_STATUS] = ""
            for k in _insee_leaf_keys:
                row_dict[k] = ""
            out_rows.append(row_dict)
            continue
        if i in invalid_indices:
            row_dict[COL_STATUS] = "INVALID_SIRET"
            row_dict[COL_SIRET] = invalid_values.get(i, row_dict.get(COL_SIRET, ""))
            for k in _insee_leaf_keys:
                row_dict[k] = ""
            out_rows.append(row_dict)
            continue
        if i not in results:
            # Should not happen if to_fetch covered all non-skipped
            row_dict[COL_STATUS] = ""
            for k in _insee_leaf_keys:
                row_dict[k] = ""
            out_rows.append(row_dict)
            continue
        st, flat = results[i]
        row_dict[COL_STATUS] = st
        if flat:
            for k in _insee_leaf_keys:
                row_dict[k] = _cell_value(flat.get(k)) if k in flat else ""
        else:
            for k in _insee_leaf_keys:
                row_dict[k] = ""
        out_rows.append(row_dict)

    out_columns = orig_cols + [COL_STATUS] + _insee_leaf_keys

    write_workbook_excel_skill(out_path, out_columns, out_rows, _insee_leaf_keys)

    if args.debug_json:
        _dump_debug_json(
            Path(args.debug_json),
            df,
            results,
            skipped_indices,
            _insee_leaf_keys,
        )

    # Summary
    n_skipped = len(skipped_indices)
    n_found = sum(1 for i in range(n_total) if i in results and results[i][0] == "OK")
    n_not_found = sum(
        1
        for i in range(n_total)
        if i in results and results[i][0] != "OK"
    )

    lines = [
        f"Total rows: {n_total}",
        f"Found (OK): {n_found}",
        f"Not found / API error: {n_not_found}",
        f"Skipped (empty FR_SIRET): {n_skipped}",
        f"Output: {out_path.resolve()}",
        f"Error log: {err_log_path.resolve()}",
    ]
    if args.debug_json:
        lines.append(f"Debug JSON: {Path(args.debug_json).resolve()}")
    print("\n".join(lines))


if __name__ == "__main__":
    main()
